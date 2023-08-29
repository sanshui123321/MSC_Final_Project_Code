import csv
import openpyxl
import pandas as pd

input_file = '/Users/sanshui/Documents/Research Project/bose_speaker/IP_DNS/IP_DNS.txt'
output_file = '/Users/sanshui/Documents/Research Project/bose_speaker/IP_DNS/IP_DNS.xlsx'

# Create a dictionary to save the list of domain names and corresponding IP addresses
domain_ip_dict = {}

# Input file
with open(input_file, 'r') as file:
    reader = csv.reader(file, delimiter=',')
    for row in reader:
        # If the row has only one element (domain name), skip the row
        if len(row) <= 1:
            continue
        domain = row[0]
        ips = row[1:]
        domain_ip_dict[domain] = ips

# Create a new Excel file
workbook = openpyxl.Workbook()
sheet = workbook.active

# Write data to Excel file
for i, (domain, ips) in enumerate(domain_ip_dict.items(), start=1):
    sheet.cell(row=i, column=1, value=domain)
    for j, ip in enumerate(ips, start=2):
        sheet.cell(row=i, column=j, value=ip)

# Save the Excel file as a temporary file
temp_file = "/tmp/temp_output.xlsx"
workbook.save(temp_file)

# Read temporary file
df = pd.read_excel(temp_file)

# Delete rows with all empty cells starting from the second column
df_cleaned = df.dropna(subset=df.columns[1:], how='all')

# Save the cleaned data to the final Excel file
df_cleaned.to_excel(output_file, index=False)
